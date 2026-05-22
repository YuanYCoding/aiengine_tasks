"""
AI引擎批量训练服务 - FastAPI 主入口（多用户会话版）
"""

import json
import asyncio
import queue
import time
import os as _os
import subprocess
import threading
import uuid
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from contextlib import asynccontextmanager

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, Response
from fastapi.middleware.cors import CORSMiddleware

from cdp_capture import CDPCaptureServer
from api_engine import AIApiClient
from batch_worker import BatchWorker


@asynccontextmanager
async def lifespan(app: FastAPI):
    asyncio.create_task(_event_pump_loop())
    asyncio.create_task(_session_cleanup_loop())
    yield


app = FastAPI(title="AI引擎批量训练服务", version="3.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = Path(__file__).parent / "static"
STATIC_DIR.mkdir(exist_ok=True)

# --- 全局单例（API客户端和批处理共享） ---
api_client: "AIApiClient | None" = None
batch_worker: "BatchWorker | None" = None
ws_clients: list[WebSocket] = []
_event_queue: queue.Queue = queue.Queue()

# --- 浏览器会话池 ---
_display_lock = threading.Lock()
_vnc_port_lock = threading.Lock()
_used_displays: set[int] = set()
_next_vnc_port = 5920
browser_sessions: dict[str, dict] = {}  # session_id → session_data


def _allocate_display() -> int:
    """分配一个可用的 Xvfb display 号"""
    with _display_lock:
        for d in range(100, 200):
            if d not in _used_displays:
                _used_displays.add(d)
                return d
        raise RuntimeError("没有可用的 display（100-199 全部占用）")


def _release_display(d: int):
    with _display_lock:
        _used_displays.discard(d)


def _allocate_vnc_port() -> int:
    global _next_vnc_port
    with _vnc_port_lock:
        port = _next_vnc_port
        _next_vnc_port += 1
        if _next_vnc_port > 5990:
            _next_vnc_port = 5920
        return port


def _start_xvfb(display: int):
    """启动 Xvfb 虚拟显示器"""
    _os.system(f"rm -f /tmp/.X{display}-lock /tmp/.X11-unix/X{display}")
    subprocess.Popen(
        ["Xvfb", f":{display}", "-screen", "0", "1920x1080x24", "-ac", "+extension", "RANDR"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    time.sleep(0.8)


def _start_x11vnc(display: int, vnc_port: int):
    """启动 x11vnc"""
    subprocess.Popen(
        ["x11vnc", "-display", f":{display}", "-nopw", "-forever", "-shared",
         "-xrandr", "-listen", "127.0.0.1", "-rfbport", str(vnc_port), "-quiet"],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    time.sleep(0.8)


def _kill_session_processes(display: int):
    """清理指定 display 的 Xvfb 和 x11vnc 进程"""
    _os.system(f"pkill -f 'Xvfb :{display}' 2>/dev/null; pkill -f 'x11vnc.*:{display}' 2>/dev/null")


async def _session_cleanup_loop():
    """定期清理超时的会话（30分钟无WebSocket连接则清理）"""
    while True:
        await asyncio.sleep(300)
        now = time.time()
        dead = []
        for sid, data in list(browser_sessions.items()):
            last_active = data.get("last_active", 0)
            cap = data.get("capture")
            # 未运行或30分钟没活跃 → 清理
            if cap and not cap.is_running and (now - last_active > 300):
                dead.append(sid)
            elif now - last_active > 1800:
                dead.append(sid)
        for sid in dead:
            await _cleanup_session(sid)
            print(f"[SESSION] cleaned up idle session {sid}")


async def _cleanup_session(sid: str):
    """清理指定会话的所有资源"""
    data = browser_sessions.pop(sid, None)
    if not data:
        return
    cap = data.get("capture")
    if cap and cap.is_running:
        cap.stop()
    display = data.get("display")
    if display is not None:
        _kill_session_processes(display)
        _release_display(display)


def get_api_client() -> AIApiClient:
    global api_client
    if api_client is None:
        api_client = AIApiClient()
    return api_client


def get_batch_worker() -> BatchWorker:
    global batch_worker
    if batch_worker is None:
        batch_worker = BatchWorker(get_api_client())
    return batch_worker


def _on_packet_captured(packet: dict):
    _event_queue.put({"type": "packet_captured", "data": packet})


async def _broadcast_ws(message: dict):
    dead = []
    for ws in ws_clients:
        try:
            await ws.send_json(message)
        except Exception:
            dead.append(ws)
    for ws in dead:
        ws_clients.remove(ws)


async def _event_pump():
    while True:
        try:
            event = _event_queue.get_nowait()
            await _broadcast_ws(event)
        except queue.Empty:
            await asyncio.sleep(0.1)


async def _event_pump_loop():
    while True:
        await _event_pump()


# ============================================================
#  页面路由
# ============================================================

@app.get("/", response_class=HTMLResponse)
async def index():
    html_path = STATIC_DIR / "index.html"
    if html_path.exists():
        return html_path.read_text(encoding="utf-8")
    return "<h1>Frontend not found.</h1>"


# ============================================================
#  WebSocket — 广播（所有用户共享同一批处理进度）
# ============================================================

@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    ws_clients.append(ws)
    try:
        while True:
            data = await ws.receive_text()
            try:
                msg = json.loads(data)
                if msg.get("action") == "ping":
                    await ws.send_json({"type": "pong"})
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        pass
    finally:
        if ws in ws_clients:
            ws_clients.remove(ws)


# ============================================================
#  浏览器捕获 API — 会话隔离
# ============================================================

@app.post("/api/proxy/start")
async def start_capture(
    session_id: str = Form(...),
    keyword: str = Form(default="trainRobotText"),
    username: str = Form(default=""),
    password: str = Form(default=""),
):
    """为指定会话启动独立的浏览器实例"""
    # 如果已有会话在运行，先停止
    existing = browser_sessions.get(session_id)
    if existing and existing.get("capture") and existing["capture"].is_running:
        return {"success": False, "message": "该会话已有浏览器在运行中，请先点击停止"}

    # 分配资源
    display = _allocate_display()
    vnc_port = _allocate_vnc_port()

    print(f"[SESSION] {session_id[:8]} → display=:{display} vnc_port={vnc_port}")

    # 启动虚拟显示器
    _start_xvfb(display)
    _start_x11vnc(display, vnc_port)

    # 设置环境并启动浏览器捕获
    _os.environ["DISPLAY"] = f":{display}"
    cap = CDPCaptureServer(
        display_num=display,
        on_packet_captured=_on_packet_captured,
    )
    cap.set_keyword(keyword)
    cap.username = username.strip()
    cap.password = password.strip()
    cap.start_async()

    # 等浏览器启动
    for _ in range(20):
        if cap.is_running:
            break
        time.sleep(0.5)

    browser_sessions[session_id] = {
        "display": display,
        "vnc_port": vnc_port,
        "capture": cap,
        "last_active": time.time(),
        "packets": [],
    }

    msg = "浏览器已启动"
    if username.strip():
        msg += "，将自动填写账号密码并登录"

    return {
        "success": True,
        "message": msg,
        "keyword": keyword,
        "auto_login": bool(username.strip()),
        "display": display,
        "vnc_port": vnc_port,
        "note": f"会话 {session_id[:8]}，display=:{display}，在浏览器画面中直接操作即可",
    }


@app.post("/api/proxy/stop")
async def stop_capture(session_id: str = Form(...)):
    """停止指定会话的浏览器"""
    data = browser_sessions.get(session_id)
    if not data:
        return {"success": False, "message": "会话不存在"}
    cap = data.get("capture")
    if not cap or not cap.is_running:
        return {"success": False, "message": "该会话没有正在运行的浏览器"}
    cap.stop()
    display = data.get("display")
    if display is not None:
        _kill_session_processes(display)
        _release_display(display)
    data["capture"] = None
    print(f"[SESSION] {session_id[:8]} stopped")
    return {"success": True, "message": "浏览器已关闭"}


@app.get("/api/proxy/status")
async def capture_status(session_id: str = ""):
    """获取指定会话的捕获状态"""
    if not session_id:
        return {"is_running": False, "keyword": "", "last_error": ""}
    data = browser_sessions.get(session_id)
    if not data or not data.get("capture"):
        return {"is_running": False, "keyword": "", "last_error": ""}
    cap = data["capture"]
    data["last_active"] = time.time()
    return {
        "is_running": cap.is_running,
        "keyword": cap.keyword_filter,
        "last_error": cap.last_error,
    }


@app.get("/api/proxy/packets")
async def get_packets(session_id: str = ""):
    """获取指定会话的捕获数据包"""
    if not session_id:
        return {"packets": []}
    data = browser_sessions.get(session_id)
    if not data or not data.get("capture"):
        return {"packets": []}
    data["last_active"] = time.time()
    return {"packets": data["capture"].get_captured_packets()}


@app.delete("/api/proxy/packets")
async def clear_packets(session_id: str = ""):
    """清空指定会话的数据包"""
    if not session_id or session_id not in browser_sessions:
        return {"success": False}
    data = browser_sessions[session_id]
    if data.get("capture"):
        data["capture"].clear_packets()
    return {"success": True}


@app.post("/api/proxy/keyword")
async def set_keyword(session_id: str = Form(...), keyword: str = Form(...)):
    """更新指定会话的捕获关键词"""
    data = browser_sessions.get(session_id)
    if data and data.get("capture"):
        data["capture"].set_keyword(keyword)
        return {"success": True, "keyword": keyword}
    return {"success": False}


# ============================================================
#  VNC WebSocket 代理 — 按会话路由到对应端口
# ============================================================

@app.websocket("/api/vnc/ws")
async def vnc_ws_proxy(ws: WebSocket):
    """WebSocket → TCP VNC 桥接（按 session_id 路由）"""
    session_id = ws.query_params.get("session_id", "")

    # 确定 VNC 端口
    data = browser_sessions.get(session_id)
    vnc_port = data["vnc_port"] if data else 5900
    if data:
        data["last_active"] = time.time()

    await ws.accept()
    try:
        reader, writer = await asyncio.open_connection("127.0.0.1", vnc_port)
    except Exception as e:
        await ws.close(code=1011, reason=str(e))
        return

    async def ws_to_tcp():
        try:
            while True:
                data = await ws.receive_bytes()
                writer.write(data)
                await writer.drain()
                if data:
                    data_session = browser_sessions.get(session_id)
                    if data_session:
                        data_session["last_active"] = time.time()
        except Exception:
            pass

    async def tcp_to_ws():
        try:
            while True:
                data = await reader.read(8192)
                if not data:
                    break
                await ws.send_bytes(data)
        except Exception:
            pass

    t1 = asyncio.create_task(ws_to_tcp())
    t2 = asyncio.create_task(tcp_to_ws())
    try:
        await asyncio.gather(t1, t2)
    except Exception:
        pass
    finally:
        t1.cancel(); t2.cancel()
        try:
            writer.close()
            await writer.wait_closed()
        except Exception:
            pass


@app.get("/vnc", response_class=HTMLResponse)
async def vnc_viewer():
    """浏览器实时画面查看器"""
    vnc_path = STATIC_DIR / "vnc.html"
    if vnc_path.exists():
        return vnc_path.read_text(encoding="utf-8")
    return "<h1>VNC viewer not found</h1>"


# ============================================================
#  API 客户端配置 API（全局共享）
# ============================================================

@app.post("/api/client/config")
async def configure_client(
    base_url: str = Form(default="https://antaios-op.100credit.com"),
    session_id: str = Form(default=""),
    voigpt_client_id: str = Form(default=""),
    cookie_string: str = Form(default=""),
):
    client = get_api_client()
    client.update_config(
        base_url=base_url,
        session_id=session_id,
        voigpt_client_id=voigpt_client_id,
        cookie_string=cookie_string,
    )
    return {"success": True, "message": "配置已更新"}


@app.get("/api/client/config")
async def get_client_config():
    client = get_api_client()
    return {
        "base_url": client.base_url,
        "session_id": client.session_id,
        "voigpt_client_id": client.voigpt_client_id,
    }


# ============================================================
#  对话测试 API
# ============================================================
import concurrent.futures
_chat_executor = concurrent.futures.ThreadPoolExecutor(max_workers=4)


@app.post("/api/chat/init")
async def chat_init(process_id: str = Form(...), variable_node: str = Form(default="{}")):
    client = get_api_client()
    print(f"[CHAT] init: process_id={process_id}")

    def _do():
        return client.create_dialog(process_id, variable_node)

    loop = asyncio.get_event_loop()
    try:
        dialog_id, opening_msg = await loop.run_in_executor(_chat_executor, _do)
        return {"success": True, "dialog_id": dialog_id, "opening_msg": opening_msg}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"success": False, "error": str(e)}


@app.post("/api/chat/send")
async def chat_send(
    process_id: str = Form(...),
    dialog_id: str = Form(...),
    text: str = Form(...),
):
    client = get_api_client()

    def _do():
        return client.single_round(text, dialog_id, process_id)

    loop = asyncio.get_event_loop()
    try:
        result = await loop.run_in_executor(_chat_executor, _do)
        return {"success": True, "ai_reply": result["ai"], "cmd": result.get("cmd", ""), "intention": result.get("intention", "")}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"success": False, "error": str(e)}


# ============================================================
#  模版下载 API
# ============================================================

@app.post("/api/template/download")
async def download_template(variable_keys: str = Form(default=""), variable_values: str = Form(default="{}")):
    import tempfile, traceback
    try:
        keys = json.loads(variable_keys) if variable_keys else []
        vals = json.loads(variable_values) if variable_values else {}
        wb = Workbook()
        ws = wb.active; ws.title = "测试集模版"
        hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center")
        headers = ["通话ID", "用户话术"] + keys
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfill; c.alignment = ha
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 2 else 22
        var_example = [vals.get(k, "") if isinstance(vals, dict) else "" for k in keys]
        example_row = ["1001", "你好，我想查询订单状态"] + var_example
        for col, v in enumerate(example_row, 1):
            ws.cell(row=2, column=col, value=v)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = tmp.name; tmp.close()
        wb.save(tmp_path)
        with open(tmp_path, "rb") as f:
            content = f.read()
        try: _os.unlink(tmp_path)
        except Exception: pass
        return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=AI_test_template.xlsx"})
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e)}


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename:
        return {"success": False, "error": "未选择文件"}
    ext = Path(file.filename).suffix.lower()
    if ext not in (".xlsx", ".xls"):
        return {"success": False, "error": f"不支持的文件格式: {ext}，仅支持 .xlsx / .xls"}
    content = await file.read()
    if not content:
        return {"success": False, "error": "文件为空"}
    try:
        worker = get_batch_worker()
        test_set = worker.parse_excel_to_test_set(content, file.filename)
        call_ids = list(dict.fromkeys(item["call_id"] for item in test_set))
        return {"success": True, "filename": file.filename, "count": len(test_set),
                "dialog_count": len(call_ids), "test_set": test_set}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ============================================================
#  批处理 API
# ============================================================

@app.post("/api/batch/run")
async def run_batch(
    process_id: str = Form(...),
    test_set: str = Form(default=""),
    variable_node: str = Form(default="{}"),
    thread_count: int = Form(default=1),
    per_dialog_messages: int = Form(default=0),
    retry_count: int = Form(default=1),
):
    try:
        test_set_list = json.loads(test_set)
    except json.JSONDecodeError:
        return {"success": False, "error": "test_set 参数格式错误"}
    if not test_set_list:
        return {"success": False, "error": "测试集为空"}
    worker = get_batch_worker()
    if worker.is_running:
        return {"success": False, "error": "已有批处理任务在运行中"}

    def safe_broadcast(msg):
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                asyncio.run_coroutine_threadsafe(_broadcast_ws(msg), loop)
        except Exception: pass

    worker.on_progress = lambda cur, tot, res, prog: safe_broadcast(
        {"type": "batch_progress", "data": {"current": cur, "total": tot, "result": res, "progress": prog}})
    worker.on_complete = lambda results: safe_broadcast(
        {"type": "batch_complete", "data": {"results": results}})
    worker.on_log = lambda msg: safe_broadcast(
        {"type": "batch_log", "data": {"message": msg}})
    worker.run_batch_async(test_set_list, process_id, variable_node, thread_count, per_dialog_messages, retry_count)
    call_ids = list(dict.fromkeys(item["call_id"] for item in test_set_list))
    return {"success": True, "message": f"批处理已启动：{len(call_ids)} 个对话，{len(test_set_list)} 条话术",
            "total": len(test_set_list), "dialog_count": len(call_ids)}


@app.post("/api/batch/stop")
async def stop_batch():
    worker = get_batch_worker()
    if not worker.is_running:
        return {"success": False, "error": "没有正在运行的批处理任务"}
    worker.stop()
    return {"success": True, "message": "批处理已停止"}


@app.get("/api/batch/status")
async def batch_status():
    return get_batch_worker().get_summary()


# ============================================================
#  导出 API
# ============================================================

@app.get("/api/export/excel")
async def export_excel():
    worker = get_batch_worker()
    if not worker.results:
        return Response(status_code=404, content="没有可导出的结果")
    excel_bytes = worker.export_excel()
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=ai_batch_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"})


@app.get("/api/export/txt")
async def export_txt():
    worker = get_batch_worker()
    if not worker.results:
        return Response(status_code=404, content="没有可导出的结果")
    txt_content = worker.export_txt()
    return Response(content=txt_content, media_type="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f"attachment; filename=ai_batch_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"})


# ============================================================
#  手动解析数据包 API
# ============================================================

def _parse_raw_http(text: str) -> dict:
    lines = text.strip().split("\n")
    if not lines: return {}
    parts = lines[0].strip().split(" ")
    method = parts[0] if len(parts) >= 1 else ""
    url_or_path = parts[1] if len(parts) >= 2 else ""
    headers = {}
    body_start = 0
    for i, line in enumerate(lines[1:], 1):
        if line.strip() == "": body_start = i + 1; break
        if ":" in line: k, v = line.split(":", 1); headers[k.strip()] = v.strip()
    body = "\n".join(lines[body_start:]).strip() if body_start > 0 else ""
    return {"method": method, "url": url_or_path, "headers": headers, "body": body}


def _parse_curl(text: str) -> dict:
    import re
    url_match = re.search(r"curl\s+['\"]([^'\"]+)['\"]", text)
    url = url_match.group(1) if url_match else ""
    headers = {}
    for m in re.finditer(r"(?:-H|--header)\s+['\"]([^'\"]+)['\"]", text):
        line = m.group(1)
        if ":" in line: k, v = line.split(":", 1); headers[k.strip()] = v.strip()
    cookie_match = re.search(r"(?:-b|--cookie)\s+['\"]([^'\"]+)['\"]", text)
    if cookie_match and "Cookie" not in headers: headers["Cookie"] = cookie_match.group(1)
    body = ""
    for flag in ("--data-raw", "--data", "-d"):
        m = re.search(rf"{flag}\s+['\"]([^'\"]+)['\"]", text)
        if m: body = m.group(1); break
    method = "POST" if body else "GET"
    return {"method": method, "url": url, "headers": headers, "body": body}


@app.post("/api/parse/manual")
async def parse_manual_packet(raw_request: str = Form(...)):
    import re
    from urllib.parse import unquote
    text = raw_request.strip()
    if not text:
        return {"success": False, "error": "请求数据为空"}
    parsed = _parse_curl(text) if text.startswith("curl ") else _parse_raw_http(text)
    headers = parsed.get("headers", {})
    body = parsed.get("body", "")
    url_or_path = parsed.get("url", "")
    cookie = headers.get("Cookie", "")
    session_id = headers.get("sessionId", "")
    voigpt_client_id = headers.get("voigpt-client-id", "")
    process_id = ""
    pid_match = re.search(r"processId[=:]\s*(\d+)", body, re.IGNORECASE)
    if not pid_match: pid_match = re.search(r"processId[=:]\s*(\d+)", url_or_path, re.IGNORECASE)
    if pid_match: process_id = pid_match.group(1)
    variable_node = "{}"
    vn_match = re.search(r"variableNode=([^&\s]+)", body, re.IGNORECASE)
    if vn_match:
        raw = vn_match.group(1)
        try: decoded = unquote(raw); json.loads(decoded); variable_node = decoded
        except Exception: variable_node = raw
    variable_keys = []
    try:
        vn_dict = json.loads(variable_node)
        if isinstance(vn_dict, dict) and vn_dict: variable_keys = list(vn_dict.keys())
    except Exception: pass
    return {"success": True, "parsed": {
        "method": parsed.get("method"), "url": url_or_path,
        "cookie": cookie, "session_id": session_id, "voigpt_client_id": voigpt_client_id,
        "process_id": process_id, "variable_node": variable_node, "variable_keys": variable_keys,
    }}


# ============================================================
#  静态文件
# ============================================================

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ============================================================
#  启动入口
# ============================================================

if __name__ == "__main__":
    import uvicorn
    _port = int(_os.environ.get("PORT", "8001"))
    uvicorn.run(app, host="0.0.0.0", port=_port, log_level="info")
