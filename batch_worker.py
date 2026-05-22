"""
批处理工作引擎
- 按通话ID分组对话，逐句调用 AI 引擎 API 获取回复
- 链路：create_dialog → 开场白 → 逐句 send_message(判断answerMsg/SSE)
"""
import io, os, re, json, time, threading
from collections import OrderedDict
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional, Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api_engine import AIApiClient


@dataclass
class BatchResult:
    index: int
    role: str = ""          # "opening" | "user" | "ai"
    content: str = ""
    call_id: str = ""       # 通话ID，标识所属对话
    dialog_id: str = ""
    cmd: str = ""
    timestamp: str = ""
    status: str = "success"
    error_msg: str = ""

    def to_dict(self):
        return asdict(self)


class BatchWorker:

    def __init__(self, api_client: AIApiClient):
        self.api_client = api_client
        self.results: list[BatchResult] = []
        self.is_running = False
        self._stop_flag = False
        self.progress = {"current": 0, "total": 0, "dialog_current": 0, "dialog_total": 0,
                         "call_id": "", "stage": ""}
        self.on_progress: Optional[Callable] = None
        self.on_complete: Optional[Callable] = None
        self.on_log: Optional[Callable] = None

    def _log(self, msg: str):
        if self.on_log:
            self.on_log(msg)
        print(f"[Batch] {msg}")

    def parse_excel_to_test_set(self, file_content: bytes, filename: str) -> list[dict]:
        """解析 Excel 文件为测试集：第1列=通话ID，第2列=用户话术"""
        ext = os.path.splitext(filename)[1].lower()
        df = pd.read_excel(io.BytesIO(file_content), engine="openpyxl" if ext == ".xlsx" else "xlrd")
        if df.empty:
            raise ValueError("Excel 文件中没有数据")
        if df.shape[1] < 2:
            raise ValueError("Excel 至少需要两列：通话ID、用户话术")
        test_set = []
        for _, row in df.iterrows():
            cid = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            text = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if cid and text:
                test_set.append({"call_id": cid, "user_text": text})
        if not test_set:
            raise ValueError("未解析到有效数据（通话ID和用户话术列均非空）")
        return test_set

    def stop(self):
        self._stop_flag = True

    def run_batch(
        self,
        test_set: list[dict],
        process_id: str,
        variable_node: str = "{}",
        thread_count: int = 1,
        per_dialog_messages: int = 0,
        retry_count: int = 1,
    ):
        """
        按通话ID分组执行批处理。
        test_set: [{"call_id": "xxx", "user_text": "yyy", ...}, ...]
        """
        self._stop_flag = False
        self.is_running = True
        self.results = []
        self._variable_node = variable_node  # 保存供导出使用

        groups = OrderedDict()
        for item in test_set:
            cid = item.get("call_id", "default")
            groups.setdefault(cid, []).append(item.get("user_text", ""))

        total_dialogs = len(groups)
        total_sentences = len(test_set)
        self.progress = {"current": 0, "total": total_sentences,
                         "dialog_current": 0, "dialog_total": total_dialogs,
                         "call_id": "", "stage": "starting"}
        seq = 0
        processed = 0

        self._log(f"开始批处理：{total_dialogs} 个对话，{total_sentences} 条话术")

        for di, (call_id, sentences) in enumerate(groups.items()):
            if self._stop_flag:
                self._log("批处理已停止")
                break

            self.progress["dialog_current"] = di + 1
            self.progress["call_id"] = call_id
            self.progress["stage"] = "init_dialog"
            self._log(f"[对话 {di+1}/{total_dialogs}] 通话ID={call_id}，{len(sentences)} 轮")

            # 初始化对话（携带变量）
            try:
                dialog_id, opening = self.api_client.create_dialog(process_id, variable_node)
                seq += 1
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.results.append(BatchResult(
                    index=seq, role="opening", content=opening or "(无开场白)",
                    call_id=call_id, dialog_id=dialog_id, timestamp=ts))
                self._broadcast_progress(processed, total_sentences, self.results[-1])
                self._log(f"  开场白: {(opening or '')[:80]}")
            except Exception as e:
                self._log(f"  初始化对话失败: {e}")
                seq += 1
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                err = BatchResult(index=seq, role="ai", content="", call_id=call_id,
                                  timestamp=ts, status="error", error_msg=f"create_dialog: {e}")
                self.results.append(err)
                self._broadcast_progress(processed, total_sentences, err)
                continue

            msg_count = 0

            for si, sentence in enumerate(sentences):
                if self._stop_flag:
                    self._log("批处理已停止")
                    break

                processed += 1
                self.progress["current"] = processed
                self.progress["stage"] = f"round {si+1}/{len(sentences)}"
                self._log(f"  [{si+1}/{len(sentences)}] 输入: {sentence}")

                # 需要在对话内部切换时重新初始化
                if per_dialog_messages > 0 and msg_count >= per_dialog_messages > 0:
                    try:
                        dialog_id, opening = self.api_client.create_dialog(process_id, variable_node)
                        msg_count = 0
                        seq += 1
                        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.results.append(BatchResult(
                            index=seq, role="opening", content=opening or "(无开场白)",
                            call_id=call_id, dialog_id=dialog_id, timestamp=ts))
                        self._broadcast_progress(processed, total_sentences, self.results[-1])
                    except Exception as e:
                        self._log(f"  重新初始化对话失败: {e}")
                        continue

                # 记录用户输入
                seq += 1
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.results.append(BatchResult(
                    index=seq, role="user", content=sentence,
                    call_id=call_id, dialog_id=dialog_id, timestamp=ts))
                self._broadcast_progress(processed, total_sentences, self.results[-1])

                # 发送消息获取AI回复（支持重试）
                ai = None; last_err = ""
                for retry_i in range(max(1, retry_count + 1)):
                    try:
                        result = self.api_client.single_round(sentence, dialog_id, process_id)
                        msg_count += 1
                        seq += 1
                        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ai = BatchResult(
                            index=seq, role="ai", content=result["ai"],
                            call_id=call_id, dialog_id=dialog_id,
                            cmd=result.get("cmd", ""), timestamp=ts, status="success")
                        self._log(f"  AI回复: {result['ai'][:100]}")
                        break
                    except Exception as e:
                        last_err = str(e)
                        if retry_i < retry_count:
                            self._log(f"  错误(重试{retry_i+1}/{retry_count}): {e}")
                            time.sleep(0.8)
                if ai is None:
                    seq += 1
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ai = BatchResult(
                        index=seq, role="ai", content="",
                        call_id=call_id, dialog_id=dialog_id,
                        timestamp=ts, status="error", error_msg=last_err)
                    self._log(f"  AI回复错误(已重试{retry_count}次): {last_err}")

                self.results.append(ai)
                self._broadcast_progress(processed, total_sentences, ai)
                time.sleep(0.5)

        self.progress["stage"] = "completed"
        self.is_running = False
        self._log(f"批处理完成：{len(self.results)} 条记录，"
                  f"{sum(1 for r in self.results if r.status=='success')} 成功，"
                  f"{sum(1 for r in self.results if r.status=='error')} 失败")

        if self.on_complete:
            self.on_complete([r.to_dict() for r in self.results])

    def _broadcast_progress(self, cur, tot, result):
        if self.on_progress:
            self.on_progress(cur, tot, result.to_dict(), dict(self.progress))

    def run_batch_async(self, test_set: list[dict], process_id: str,
                        variable_node: str = "{}", thread_count: int = 1, per_dialog_messages: int = 0,
                        retry_count: int = 1):
        t = threading.Thread(target=self.run_batch,
                             args=(test_set, process_id, variable_node, thread_count, per_dialog_messages, retry_count),
                             daemon=True)
        t.start()
        return t

    def _build_conversation_record(self, call_id: str) -> tuple:
        """构建单个通话ID的对话记录、请求参数、变量、用户输入、Agent回复"""
        lines = []; dialog_ids = set(); user_lines = []; agent_lines = []
        ui = 0; ai = 0
        for r in self.results:
            if r.call_id != call_id: continue
            if r.role == "opening":
                lines.append(f"Agent：{r.content}")
                dialog_ids.add(r.dialog_id)
                ai += 1; agent_lines.append(f"Agent-{ai}：{r.content}")
            elif r.role == "user":
                lines.append(f"User：{r.content}")
                ui += 1; user_lines.append(f"User-{ui}：{r.content}")
            elif r.role == "ai":
                lines.append(f"Agent：{r.content}")
                ai += 1; agent_lines.append(f"Agent-{ai}：{r.content}")
        record = "\n\n".join(lines)
        params = ", ".join(sorted(dialog_ids))
        var_str = self._format_variable()
        user_str = "\n\n".join(user_lines)
        agent_str = "\n\n".join(agent_lines)
        return record, params, var_str, user_str, agent_str

    def _format_variable(self) -> str:
        """格式化变量字典为字符串（双换行分隔）"""
        vn = getattr(self, '_variable_node', '{}')
        try:
            d = json.loads(vn) if isinstance(vn, str) else vn
            if isinstance(d, dict) and d:
                return "\n\n".join(f"{k}：{v}" for k, v in d.items())
        except (json.JSONDecodeError, Exception):
            pass
        return ""

    def export_excel(self) -> bytes:
        """导出：通话ID | 对话记录 | 变量 | 用户输入 | Agent回复 | 请求参数"""
        wb = Workbook(); ws = wb.active; ws.title = "AI对话记录"
        hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        has_vars = bool(self._format_variable())
        headers = ["通话ID", "对话记录"]
        if has_vars: headers.append("变量")
        headers += ["用户输入", "Agent回复", "请求参数"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

        seen = set(); ordered_cids = []
        for r in self.results:
            if r.call_id and r.call_id not in seen:
                seen.add(r.call_id); ordered_cids.append(r.call_id)

        for row, cid in enumerate(ordered_cids, 2):
            record, params, var_str, user_str, agent_str = self._build_conversation_record(cid)
            vals = [cid, record]
            if has_vars: vals.append(var_str)
            vals += [user_str, agent_str, params]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = Alignment(vertical="top", wrap_text=True); c.border = tb

        from openpyxl.utils import get_column_letter
        widths = [16, 70]
        if has_vars: widths.append(22)
        widths += [35, 35, 38]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for row in range(2, len(ordered_cids) + 2):
            ws.row_dimensions[row].height = 220
        out = io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()

    def export_txt(self) -> str:
        lines = ["=" * 60, f"AI引擎对话记录 - {datetime.now():%Y-%m-%d %H:%M:%S}", "=" * 60, ""]
        seen = set(); ordered_cids = []
        for r in self.results:
            if r.call_id and r.call_id not in seen:
                seen.add(r.call_id); ordered_cids.append(r.call_id)
        for cid in ordered_cids:
            record, params, var_str, user_str, agent_str = self._build_conversation_record(cid)
            lines.append(f"===== 通话ID: {cid} =====")
            lines.append(f"请求参数: {params}")
            if var_str: lines.append(f"变量:\n{var_str}")
            lines.append(f"用户输入:\n{user_str}")
            lines.append(f"Agent回复:\n{agent_str}")
            lines.append(record); lines.append("")
        return "\n".join(lines)

    def get_summary(self) -> dict:
        total = len(self.results)
        success = sum(1 for r in self.results if r.status == "success")
        error = sum(1 for r in self.results if r.status == "error")
        dialogs = len(set(r.call_id for r in self.results if r.call_id))
        return {"total": total, "success": success, "error": error,
                "dialogs": dialogs, "is_running": self.is_running,
                "progress": self.progress,
                "results": [r.to_dict() for r in self.results]}
